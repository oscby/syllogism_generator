import os
import random
import sys
import openpyxl

# INPUT
path = ''
num_of_sylls = 100

# modda nedan
subjekt = ['ateister', 'sossar']
verb = 'är'
adjektiv = ['intelligenta', 'fula', 'giriga', 'luriga', 'söta', 'förtroendeingivande', 'seriösa', 'konspiratoriska']
objekt = ['pålästa', 'olyckliga', 'pengakåta', 'manipulativa', 'jättefab', 'tillförlitliga', 'respektabla', 'otillförlitliga']
    
def modus(sub, verb, adj, obj):
    foo = random.randint(0,1)
    bar = random.randint(0,1)
    if foo == 0:
        if bar == 1:
            syll = [f'om {sub} {verb} {adj}, \nså {verb} {sub} {obj}.\n{sub} {verb} {adj}, \nalltså {verb} {sub} {obj}.', 'TRUE', 'modus_ponens']
        else:
            syll = [f'om {sub} {verb} {adj}, \nså {verb} {sub} {obj}.\n{sub} {verb} {obj}, \nalltså {verb} {sub} {adj}.', 'FALSE', 'modus_ponens']
    elif foo == 1:
        if bar == 1:
            syll = [f'om {sub} {verb} {adj}, \nså {verb} {sub} {obj}.\n{sub} {verb} inte {obj}, \nalltså {verb} inte {sub} {adj}.', 'TRUE', 'modus_tollens']
        else:
            syll = [f'om {sub} {verb} {adj}, \nså {verb} {sub} {obj}.\n{sub} {verb} inte {obj}, \nalltså {verb} {sub} {adj}.', 'FALSE', 'modus_tollens']
    return syll

def main():

    clear = lambda: os.system('clear')
    clear()

    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        wb.save(path)

    ws = wb.active

    ws.auto_filter.ref = f"A1:C{num_of_sylls+1}"
    ws.auto_filter.add_filter_column(1, ['TRUE', 'FALSE'])
    ws.auto_filter.add_filter_column(2, ['modus_ponens', 'modus_tollens'])
    ws.auto_filter.add_sort_condition(f'A2:A{num_of_sylls+1}')
    sh = wb['Sheet']

    for i in range(0, num_of_sylls):
        if i == 0:
            syll = ['Syllogism', 'Bool', 'Statement type']
        else:
            sub = subjekt[random.randint(0,len(subjekt)-1)]
            adj = adjektiv[random.randint(0,len(adjektiv)-1)]
            obj = objekt[adjektiv.index(adj)]
            syll = modus(sub, verb, adj, obj)
        sh.cell(i + 1, 1).value = syll[0]
        sh.cell(i + 1, 2).value = syll[1]
        sh.cell(i + 1, 3).value = syll[2]

        ws.row_dimensions[i+1].height = 60
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['C'].width = 20

    wb.save(path)

    sys.exit()

if __name__ == '__main__':
    main()
